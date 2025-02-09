import os
import pandas as pd
import sys
from openpyxl import load_workbook

# Define file path base directory and Excel output path
base_dir = r'C:\Users\msalman\Desktop\OSMOSE ETs\Glass\results\Glass\run_015_Uncertainity'
output_excel_path = r'C:\Users\msalman\Desktop\OSMOSE ETs\Python work\uncertainity\uncertainity_results_100125.xlsx'

plant_capacity = 292000  # t/year
scenario_name = "Scenario_2050"  # Set the scenario name directly or dynamically via user input if needed

# Read the case number from the command-line arguments
if len(sys.argv) < 2:
    raise ValueError("Please provide a case number as a command-line argument!")
case_number = int(sys.argv[1])

# Assign column names based on case number
case_names = {
    1: "NG", 2: "NG_CC", 3: "NGOxy", 4: "NGOxy_CC", 5: "Hyb",
    6: "Hyb_CC", 7: "EL", 8: "EL_CC", 9: "H2", 10: "H2_CC"
}

# Determine the case column name based on the case number
if case_number not in case_names:
    raise ValueError(f"Invalid case number! Choose from {list(case_names.keys())}.")
case_column_name = case_names[case_number]

# Create an empty list to store extracted data
data = []

# Iterate through the file range
for i in range(1, 183):  # Adjust range as needed
    file_path = os.path.join(base_dir, f's_{i:03}', 'opt', 'hc_0000.txt')
    
    # Check if the file exists before attempting to read it
    if os.path.exists(file_path):
        with open(file_path, 'r') as file:
            for line in file:
                if "KPI_totalcost" in line:
                    # Extract the value after the "=" sign and strip whitespace
                    total_cost = float(line.split('=')[1].strip()) / plant_capacity
                    data.append({'Run Number': i, case_column_name: total_cost})
                    break
    else:
        print(f"File not found: {file_path}")

# Create a pandas DataFrame from the extracted data
df = pd.DataFrame(data)

# Check if the Excel file already exists
if os.path.exists(output_excel_path):
    # Load existing workbook
    book = load_workbook(output_excel_path)
    
    if scenario_name in book.sheetnames:
        # Load the existing scenario sheet
        existing_df = pd.read_excel(output_excel_path, sheet_name=scenario_name)
        
        # Merge the new results with the existing DataFrame
        if "Run Number" in existing_df.columns:
            df = pd.merge(existing_df, df, on="Run Number", how="outer")
        else:
            # If "Run Number" doesn't exist in the sheet, create a new merged DataFrame
            df = pd.concat([existing_df, df], axis=1)
    else:
        # If the scenario sheet doesn't exist, create it
        existing_df = pd.DataFrame()
else:
    # If the Excel file doesn't exist, create a new DataFrame
    existing_df = pd.DataFrame()

# Save the combined DataFrame back to the Excel file
with pd.ExcelWriter(output_excel_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
    df.to_excel(writer, sheet_name=scenario_name, index=False)

print(f"Data extraction complete. Results saved to {output_excel_path} in the '{scenario_name}' sheet.")
