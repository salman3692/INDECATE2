import re
import pandas as pd
from openpyxl import load_workbook
import sys

# Read the case number from command line argument
case_number = int(sys.argv[1])
# case_number = 3

# Assign sheet name based on case number
case_names = {
    1: "NG", 2: "NG_CC", 3: "NGOxy", 4: "NGOxy_CC", 5: "Hyb",
    6: "Hyb_CC", 7: "EL", 8: "EL_CC", 9: "H2", 10:"H2_CC"
}

# Get the sheet name for the case number
sheet_name = case_names.get(case_number)
if not sheet_name:
    print("Invalid case number.")
    sys.exit(1)

plant_capacity = 33333.33 #kg/hr
plant_capacity_an = plant_capacity * 8.76 #kg/hr

# Define the patterns and corresponding line descriptions
patterns_and_descriptions = [
    # Investment Costs
    ("DefaultInvCost c1_Glass_NG_Furnace_NG_Furnace_Cinv                  c1_Glass_NG_Furnace_NG_Furnace", "NG Furnace Cinv"),
    ("DefaultInvCost c1_Glass_H2_Furnace_H2_Furnace_Cinv                  c1_Glass_H2_Furnace_H2_Furnace","H2 Furnace Cinv"),
    ("DefaultInvCost c1_Glass_EL_Furnace_el_Furnace_Cinv                  c1_Glass_EL_Furnace_el_Furnace","EL furnace Cinv"),
    ("DefaultInvCost c1_Glass_NGOxy_Furnace_NGOxy_Furnace_Cinv            c1_Glass_NGOxy_Furnace_NGOxy_Furnace", "NG Oxy Cinv"),
    ("DefaultInvCost c1_Glass_Hybrid_furnace_Hybrid_Furnace_Cinv          c1_Glass_Hybrid_furnace_Hybrid_Furnace", "Hyb Cinv"),
    ("DefaultInvCost c1_Glass_glass_float_glass_float_Cinv                c1_Glass_glass_float_glass_float", "Flat_glass Cinv"),
    ("DefaultInvCost c1_Glass_ORC_CO2_super_sCO2_supercycle_Cinv          c1_Glass_ORC_CO2_super_sCO2_supercycle", "ORC Cinv"),
    ("DefaultInvCost c1_Glass_CC_Units_CCSPost_Cinv                       c1_Glass_CC_Units_CCSPost", "CCS Cinv"),
    ("DefaultInvCost c1_Glass_BoilerCCS_Boiler_NG_Cinv                    c1_Glass_BoilerCCS_Boiler_NG", "Boiler Cinv"),
    ("DefaultInvCost c1_Glass_CC_Units_CPU4Oxy_Cinv                       c1_Glass_CC_Units_CPU4Oxy", "CPU Cinv"),
    ("DefaultInvCost c1_Glass_Air_separation_ASU_Cinv                     c1_Glass_Air_separation_ASU", "ASU Cinv"),
    ("DefaultOpCost  c1_world_mergedresource_dummy                        c1_world_mergedresource_Resource_Dummy","Dummy Cinv"),
    # Operating Costs
    ("DefaultOpCost  c1_world_mergedresource_Resource_Electricity_Cost    c1_world_mergedresource_Resource_Electricity", "Elec op"),
    ("DefaultOpCost  c1_world_mergedresource_Resource_Hydrogen_Cost       c1_world_mergedresource_Resource_Hydrogen", "H2 op"),
    ("DefaultOpCost  c1_world_mergedresource_Resource_Naturalgas_Cost     c1_world_mergedresource_Resource_Naturalgas", "NG op"),
    ("DefaultOpCost  c1_world_mergedwaste_Environ_Cost                    c1_world_mergedwaste_Environ", "CO2 op"),
    ("DefaultOpCost  c1_world_mergedresource_dummy                        c1_world_mergedresource_Resource_Dummy","Dummy op"),
    # CO2 Transport Costs
    ("DefaultOpCost  c1_world_mergedwaste_CO2_transport_Cost              c1_world_mergedwaste_CO2_transport", "CO2transport cost"),
    ("DefaultOpCost  c1_world_mergedresource_dummy                        c1_world_mergedresource_Resource_Dummy","Dummy cost"),
    # Energy Demand and supply
    ("layers_Electricity         c1_Glass_CC_Units_CCSPost                 1","CCS_EL demand"),
    ("layers_Electricity         c1_Glass_glass_float_glass_float          1","Furnace_EL demand"),
    ("layers_Electricity         c1_Glass_Air_separation_ASU               1","ASU_EL demand"),
    ("layers_Electricity         c1_Glass_CC_Units_CPU4Oxy                 1","CPU_EL demand"),
    ("layers_Electricity         c1_Glass_Hybrid_furnace_Hybrid_Furnace    1","HybFurnace_EL demand"),
    ("layers_Electricity         c1_Glass_EL_Furnace_el_Furnace            1","ELFurnace_EL demand"),
    ("layers_Naturalgas          c1_Glass_BoilerCCS_Boiler_NG              1","Boiler_NG demand"),
    ("layers_Naturalgas          c1_Glass_NG_Furnace_NG_Furnace            1","NGFurnace_NG demand"),
    ("layers_Naturalgas          c1_Glass_NGOxy_Furnace_NGOxy_Furnace      1","NGOxyFurnace_NG demand"),
    ("layers_Naturalgas          c1_Glass_Hybrid_furnace_Hybrid_Furnace    1","HybFurnace_NG demand"),
    ("layers_hydrogen            c1_Glass_glass_float_glass_melting_solids 1","Process_H2 demand"),
    ("layers_hydrogen            c1_Glass_H2_Furnace_H2_Furnace            1","H2furnace_H2 demand"),
    ("layers_dummy               c1_world_mergedresource_Resource_Hydrogen       1","Dummy demand"),
    ("layers_Electricity         c1_Glass_ORC_CO2_super_sCO2_supercycle          1","Elec supply"),
    ("layers_dummy               c1_world_mergedresource_Resource_Hydrogen       1","Dummy supply"),
    # Emissions Breakdown
    ("layers_EnvCO2Em            c1_world_mergedwaste_Environ              1","Scope1 Emissions"),
    ("layers_IndEnvCO2Em         c1_world_mergedwaste_Ind_Emiss            2","Scope2 Emissions"),
    ("layers_dummy               c1_world_mergedresource_Resource_dummy       1","Dummy Emissions"),
    ("layers_IndEnvCO2Em         c1_world_mergedresource_Resource_Electricity    2","EL Scope2"),
    ("layers_IndEnvCO2Em         c1_world_mergedresource_Resource_Hydrogen       2","H2 Scope2"),
    ("layers_IndEnvCO2Em         c1_world_mergedresource_Resource_Naturalgas     2","NG Scope2"),
    ("layers_dummy               c1_world_mergedresource_Resource_dummy     1","Dummy Scope2"),
    ("KPI_impact =","total impact")
]

# sheet_name = 'NG'

# Define input file paths with descriptive variable names
path_2024 = r'C:\Users\msalman\Desktop\OSMOSE ETs\Glass\results\Glass\run_005_testing\s_001\opt\hc_0000.txt'
path_2030 = r'C:\Users\msalman\Desktop\OSMOSE ETs\Glass\results\Glass\run_005_testing\s_002\opt\hc_0000.txt'
path_2040 = r'C:\Users\msalman\Desktop\OSMOSE ETs\Glass\results\Glass\run_005_testing\s_003\opt\hc_0000.txt'
path_2050 = r'C:\Users\msalman\Desktop\OSMOSE ETs\Glass\results\Glass\run_005_testing\s_004\opt\hc_0000.txt'

# Define output file path
output_file_path = r'C:\Users\msalman\Desktop\OSMOSE ETs\Python work\INDECATE2\data\Results_Scenarios_TESTING.xlsx'

# Load the existing workbook
workbook = load_workbook(output_file_path)

# Select the specified sheet or create it if it doesn't exist
if sheet_name in workbook.sheetnames:
    worksheet = workbook[sheet_name]
    # Clear existing data in the sheet
    worksheet.delete_rows(1, worksheet.max_row)  # Delete all rows in the sheet
else:
    worksheet = workbook.create_sheet(title=sheet_name)

# Create a dictionary to store extracted data
extracted_data = {description: [] for _, description in patterns_and_descriptions}


# ##############################################################################################################################""
# Process each input file
for path_name, input_file_path in [('path_2024', path_2024), ('path_2030', path_2030), ('path_2040', path_2040), ('path_2050', path_2050)]:
    with open(input_file_path, 'r') as file:
        lines = file.readlines()

        for line in lines:
            for pattern, description in patterns_and_descriptions:
                if pattern in line:
                    # Extract the number value using regular expression
                    value_match = re.search(r'\d+(\.\d+)?$', line)
                    if value_match:
                        value = float(value_match.group())
                        extracted_data[description].append(value)


# Write header row
header_row = ['Lines', 'path_2024', 'path_2030', 'path_2040', 'path_2050']
worksheet.append(header_row)

# Write data rows
for pattern, description in patterns_and_descriptions:
    data_row = [description]
    for path_name in ['path_2024', 'path_2030', 'path_2040', 'path_2050']:
        if extracted_data[description]:
            data_row.append(extracted_data[description].pop(0))
        else:
            data_row.append('')  # Add an empty cell for missing values
    worksheet.append(data_row)


# Calculate and append the sum row
cinv_sum_row = ['Spec Cinv']
op_sum_row = ['Spec Op']
CO2TrCost_row = ['Spec CO2TrCost']
energy_sum_row = ['Spec Energy']
HREC_row = ['Spec HREC']
emissions_row = ['Spec Emissions']

# Initialize column sums
column_sums = [0.0, 0.0, 0.0, 0.0]
CO2TrCost_columns = [0.0, 0.0, 0.0, 0.0]  # Separate for CO2 Transport Cost
energy_column_sums = [0.0, 0.0, 0.0, 0.0]  # Separate sums for energy
HREC_columns = [0.0, 0.0, 0.0, 0.0]  # Separate for HREC
emissions_column = [0.0, 0.0, 0.0, 0.0]  # Separate for emissions

for path_name in ['path_2024', 'path_2030', 'path_2040', 'path_2050']:
    column_index = ['path_2024', 'path_2030', 'path_2040', 'path_2050'].index(path_name) + 2  # Calculate the integer index for the column
    column_values = [data for data in worksheet.iter_cols(min_col=column_index, max_col=column_index, values_only=True)][0]

    cinv_values = []
    op_values = []
    CO2TrCost_value = []
    energy_values = []
    HREC_value = []
    emissions_values = []

    for i, description in enumerate([data[0] for data in worksheet.iter_rows(values_only=True)][1:]):
        value = column_values[i]
        
        if description is None:
            continue  # Skip this iteration if the description is None
        
        if isinstance(description, str) and description.lower().endswith('cinv'):
            if isinstance(value, (int, float)):
                cinv_values.append(value)
        elif isinstance(description, str) and description.lower().endswith('op'):
            if isinstance(value, (int, float)):
                op_values.append(value)
        elif isinstance(description, str) and description.lower().endswith('cost'):
            if isinstance(value, (int, float)):
                CO2TrCost_value.append(value)
        elif isinstance(description, str) and description.lower().endswith('demand'):
            if isinstance(value, (int, float)):
                energy_values.append(value)
        elif isinstance(description, str) and description.lower().endswith('emissions'):
            if isinstance(value, (int, float)):
                emissions_values.append(value)
        elif isinstance(description, str) and description.lower().endswith('supply'):
            if isinstance(value, (int, float)):
                HREC_value.append(value)

    cinv_sum = sum(cinv_values) / plant_capacity_an if cinv_values else 0.0

    FixOp_sum = 25201727  # EUR/yr
    op_sum = (sum(op_values) + FixOp_sum) / plant_capacity_an if op_values else 0.0

    CO2TrCost = CO2TrCost_value[0] / plant_capacity_an if CO2TrCost_value else 0.0

    energy_sum = sum(energy_values) / plant_capacity if energy_values else 0.0

    HREC = HREC_value[0] / plant_capacity if HREC_value else 0.0

    emissions_sum = sum(emissions_values) / plant_capacity if emissions_values else 0.0

    cinv_sum_row.append(cinv_sum)
    op_sum_row.append(op_sum)
    CO2TrCost_row.append(CO2TrCost)
    energy_sum_row.append(energy_sum)
    HREC_row.append(HREC)
    emissions_row.append(emissions_sum)

    # Update column sums
    column_sums[column_index - 2] += cinv_sum + op_sum + CO2TrCost
    energy_column_sums[column_index - 2] += energy_sum - HREC
    emissions_column[column_index - 2] += emissions_sum

# Add a row to sum "Cinv" and "Op" values
total_sum_row = ['Total Spec Cost (EUR/t)']
total_sum_row.extend(column_sums)

# Append the "Energy" sum row separately
worksheet.append(energy_sum_row)
worksheet.append(HREC_row)
worksheet.append(emissions_row)
worksheet.append(cinv_sum_row)
worksheet.append(op_sum_row)
worksheet.append(CO2TrCost_row)
worksheet.append(total_sum_row)

# Save the workbook to the specified output file
workbook.save(output_file_path)
print(f"Data successfully written to '{output_file_path}' under the sheet '{sheet_name}'.")


##################################################################################################################